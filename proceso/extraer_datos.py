#!/usr/bin/env python3
"""Extrae y filtra los datos de la semana a _data.json.
Uso: python3 proceso/extraer_datos.py <carpeta_datos> <dd-mm-yyyy>
Por defecto: Ejecuciones/<hoy>/datos"""
import openpyxl, json, datetime, glob, os, sys
from collections import Counter

FECHA = sys.argv[2] if len(sys.argv) > 2 else datetime.date.today().strftime("%d-%m-%Y")
DDIR = sys.argv[1] if len(sys.argv) > 1 else f"Ejecuciones/{FECHA}/datos"
HOY = datetime.datetime.strptime(FECHA, "%d-%m-%Y")

ESTADO_EXCLUIR = {"Cerrado", "Cancelado"}          # <-- filtro de estado
TIPO_INCLUIR = None                                 # None = todos (INC + RITM)

def fmt(v): return v.strftime("%d-%m-%Y %H:%M") if isinstance(v, datetime.datetime) else ("" if v is None else str(v))
def fdate(v): return v.strftime("%d-%m-%Y") if isinstance(v, datetime.datetime) else ("" if v is None else str(v))

tick_file = des_file = None
for f in glob.glob(DDIR + "/*.xlsx"):
    wb = openpyxl.load_workbook(f, read_only=True)
    if "Page 1" in wb.sheetnames: tick_file = f
    if "Desarrollos" in wb.sheetnames: des_file = f
    wb.close()

ws = openpyxl.load_workbook(tick_file, data_only=True)["Page 1"]
tickets = []
for r in range(2, ws.max_row + 1):
    est, grp = ws.cell(r, 3).value, ws.cell(r, 9).value
    if est is None or grp is None: continue
    if str(est).strip() in ESTADO_EXCLUIR: continue
    creado = ws.cell(r, 7).value
    dias = (HOY - creado).days if isinstance(creado, datetime.datetime) else ""
    tickets.append({"estado": str(est).strip(), "sociedad": fmt(ws.cell(r, 2).value),
        "numero": fmt(ws.cell(r, 1).value), "resumen": fmt(ws.cell(r, 4).value),
        "apertura": fmt(creado), "usuario": fmt(ws.cell(r, 5).value), "dias": dias,
        "_grupo": str(grp).strip()})

ws2 = openpyxl.load_workbook(des_file, data_only=True)["Desarrollos"]
def col(name):
    for c in range(1, ws2.max_column + 1):
        if str(ws2.cell(1, c).value).strip().rstrip('.').lower() == name: return c
cA, cB, cC, cG, cH, cK, cL, cT, cW = (col('codigo'), col('tipo de requerimiento'), col('título'),
    col('status'), col('frente'), col('creado'), col('solicitante'),
    col('fecha plan entrega qa'), col('fecha entrega a qa'))
desarrollos = []
for r in range(2, ws2.max_row + 1):
    fr = ws2.cell(r, cH).value
    if fr is None: continue
    desarrollos.append({"status": str(ws2.cell(r, cG).value).strip(), "creado": fmt(ws2.cell(r, cK).value),
        "codigo": fmt(ws2.cell(r, cA).value), "tipo": fmt(ws2.cell(r, cB).value), "titulo": fmt(ws2.cell(r, cC).value),
        "plan_qa": fdate(ws2.cell(r, cT).value), "entrega_qa": fdate(ws2.cell(r, cW).value),
        "solicitante": fmt(ws2.cell(r, cL).value), "_frente": str(fr).strip()})

modulos = [
    ("comercial", "NET", "Net - Metrogas", "Portales y autoatencion"),
    ("comercial", "Sharepoint", "Sharepoint - Metrogas", None),
    ("comercial", "B&I", "B&I - Metrogas", "SAP ISU"),
    ("comercial", "FICA", "Fica - Metrogas", "SAP FICA"),
    ("comercial", "CRM", "Crm - Metrogas", "SAP CRM"),
    ("comercial", "PM", "Pm - Metrogas", "SAP PM"),
    ("comercial", "PS", "Ps - Metrogas", "SAP PS"),
    ("corporativa", "HCM", "Hcm - Metrogas", "SAP HCM"),
    ("corporativa", "SSFF", "Ssff - Metrogas", None),
    ("corporativa", "R&P", "R&P - Metrogas", None),
    ("corporativa", "SD", "Sd - Metrogas", "SAP SD"),
    ("corporativa", "BW", "Bw-Hana - Metrogas", "BI"),
    ("corporativa", "FICO", "Fico - Metrogas", "SAP FI"),
    ("corporativa", "MM", "Mm - Metrogas", "SAP MM"),
]
out = {"fecha": FECHA, "decks": {
    "comercial": {"titulo": "Torre Comercial y Técnica", "modulos": []},
    "corporativa": {"titulo": "Torre Corporativa", "modulos": []}}}
for deck, nombre, grupo, frente in modulos:
    t = [x for x in tickets if x["_grupo"] == grupo]
    t.sort(key=lambda x: x["estado"])
    d = [x for x in desarrollos if frente and x["_frente"] == frente]
    cc = Counter(x["status"] for x in d)
    # resumen = TODOS los estados presentes (orden por cantidad desc, luego nombre) + total
    items = sorted(cc.items(), key=lambda kv: (-kv[1], kv[0]))
    resumen = items + [("Total general", sum(cc.values()))]
    out["decks"][deck]["modulos"].append({"nombre": nombre,
        "tickets": [{k: v for k, v in x.items() if not k.startswith("_")} for x in t], "tickets_total": len(t),
        "desarrollos": [{k: v for k, v in x.items() if not k.startswith("_")} for x in d], "resumen": resumen})

json.dump(out, open("proceso/_data.json", "w"), ensure_ascii=False, indent=1)
for deck in ["comercial", "corporativa"]:
    tot = sum(m['tickets_total'] for m in out['decks'][deck]['modulos'])
    print(f"[{deck}] total tickets: {tot}")
print("OK -> proceso/_data.json")
